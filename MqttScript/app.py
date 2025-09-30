from flask import Flask
import paho.mqtt.client as mqtt
import time
import threading
from openpyxl import Workbook, load_workbook
from datetime import datetime
import signal
import sys
import os

# Initialize Flask app
app = Flask(__name__)

# MQTT settings
broker = "203.109.124.70"
port = 18889
topic_publish = "tubeF0BF27/control"
topic_subscribe = "tubeF0BF27/#"

# Excel file path
file_path = "mqtt_responses.xlsx"

# Check if file exists; if not, create it with headers
if not os.path.exists(file_path):
    wb = Workbook()
    ws = wb.active
    ws.append([
        'Timestamp', 'Device_ID', 'Device_Type', 'Intensity', 
        'Load Status', 'Power', 'Temperature', 'AutoBrightness Status', 
        'autoMotion Status', 'Lux', 'PIR Data', 'Required Intensity'
    ])
    wb.save(file_path)
else:
    wb = load_workbook(file_path)
    ws = wb.active  # Open existing sheet

# Event to stop threads
stop_event = threading.Event()
mqttc = mqtt.Client()

# Create a lock to manage thread-safe file access
file_lock = threading.Lock()

def on_connect(client, userdata, flags, rc):
    print("Connected with result code " + str(rc))
    client.subscribe(topic_subscribe)

def on_message(client, userdata, msg):
    response = msg.payload.decode('utf-8').strip()
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    print(f"\n📡 Raw MQTT Data: {response}")  # Debugging statement

    # Remove {} brackets and split by ':'
    data_parts = response.strip("{}").split(":")
    print(f"🔍 Parsed Data: {data_parts}")  # Debugging statement

    # Ensure we have exactly 11 values
    if len(data_parts) == 11:
        (device_id, device_type, device_status, intensity, power, temperature, 
         auto_motion_status, auto_brightness_status, voltage, current, lux) = data_parts

        # Locking file writing to avoid race conditions
        with file_lock:
            ws.append([
                timestamp, device_id, device_type, device_status, 
                auto_motion_status, intensity, power, temperature, 
                auto_brightness_status, voltage, current, lux
            ])
            wb.save(file_path)

        print(f"✅ Data Saved to Excel: {timestamp}, {device_id}, {device_type}, {device_status}, {auto_motion_status}, {intensity}, {power}, {temperature}, {auto_brightness_status}, {voltage}, {current}, {lux}")
    else:
        print("❌ Received unexpected data format:", response)

def publish_data():
    while not stop_event.is_set():  # Check for the stop event
        try:
            message = "200"
            mqttc.publish(topic_publish, message)
            print(f"📤 Published: {message}")
            time.sleep(10)
        except KeyboardInterrupt:
            print("Publishing stopped.")
            break

def start_mqtt():
    mqttc.on_connect = on_connect
    mqttc.on_message = on_message
    mqttc.connect(broker, port, 60)  # Connect to the broker
    mqttc.loop_start()  # Start the MQTT loop to handle incoming messages

@app.route('/')
def index():
    return "MQTT Publisher and Flask Server Running!"

def graceful_shutdown(signal, frame):
    print("\n⚠️  Shutting down...")
    stop_event.set()  # Set the stop event to signal threads to stop
    mqttc.loop_stop()  # Stop MQTT loop
    sys.exit(0)  # Exit the program

# Capture Ctrl+C for graceful shutdown
signal.signal(signal.SIGINT, graceful_shutdown)

if __name__ == "__main__":
    # Start the MQTT client and data publishing threads
    threading.Thread(target=start_mqtt).start()
    threading.Thread(target=publish_data).start()
    app.run(debug=True, use_reloader=False)