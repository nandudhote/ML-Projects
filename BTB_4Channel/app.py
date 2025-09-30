import paho.mqtt.client as mqtt
import sqlite3
from datetime import datetime
import ssl
import time
import os
from openpyxl import Workbook, load_workbook

# -------------------
# AWS IoT Core settings
# -------------------
broker = "a6kvi1np2cmrt-ats.iot.ap-south-1.amazonaws.com"  # <-- Replace with your AWS endpoint
port = 8883

client_id = "basicPubSub1"
ca_cert = "cert.pem"
client_cert = "test_lte.cert.pem"
client_key = "test_lte.private.key"

# Create MQTT client
mqttc = mqtt.Client(client_id="mqttx_65456344_x", protocol=mqtt.MQTTv311)  # client_id must be unique
mqttc.tls_set(ca_certs=ca_cert, certfile=client_cert, keyfile=client_key)

# -------------------
# SQLite DB setup
# -------------------
DB_FILE = "device_data.db"

def init_db():
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS four_channel (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            device_type TEXT,
            device_id TEXT,
            voltage REAL,
            current REAL,
            power REAL,
            inserttimestamp TEXT,
            relay1 INTEGER,
            relay2 INTEGER,
            relay3 INTEGER,
            relay4 INTEGER,
            voltage2 REAL,
            current2 REAL,
            power2 REAL
        );
    """)
    conn.commit()
    conn.close()

# -------------------
# Excel setup
# -------------------
EXCEL_FILE = "device_data.xlsx"

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "DeviceData"
        ws.append([
            "Device Type", "Device ID", "Voltage1", "Current1", "Power1",
            "Voltage2", "Current2", "Power2",
            "Relay1", "Relay2", "Relay3", "Relay4",
            "Insert Timestamp"
        ])
        wb.save(EXCEL_FILE)

def insert_excel_row(data):
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "DeviceData"
        ws.append([
            "Device Type", "Device ID", "Voltage1", "Current1", "Power1",
            "Voltage2", "Current2", "Power2",
            "Relay1", "Relay2", "Relay3", "Relay4",
            "Insert Timestamp"
        ])
    ws.append(data)
    wb.save(EXCEL_FILE)

# -------------------
# MQTT Callbacks
# -------------------
def on_connect(client, userdata, flags, rc):
    if rc == 0:
        print("✅ Connected to AWS IoT Core")
        client.subscribe("#")  # Subscribe to all topics
    else:
        print(f"❌ Failed to connect with return code {rc}")

def on_disconnect(client, userdata, rc):
    if rc != 0:
        print("⚠️ Disconnected from broker. Reconnecting...")
        try:
            mqttc.reconnect()
        except:
            pass

def on_message(client, userdata, message):
    try:
        payload = message.payload.decode()
        topic = message.topic
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:00')

        # Connect to SQLite DB
        connection = sqlite3.connect(DB_FILE)
        cursor = connection.cursor()

        # Only process BTB4Channel device topics
        if "BTB4Channel" in topic:
            payload_str = payload.strip("{}")
            split_payload = payload_str.split(":")

            # Expected format:
            # {device_id:DeviceId:V1:I1:P1:V2:I2:P2:R1:R2:R3:R4}
            if len(split_payload) >= 12 and split_payload[0] == "device_id":
                device_id   = split_payload[1]
                voltage1    = float(split_payload[2])
                current1    = float(split_payload[3])
                power1      = float(split_payload[4])
                voltage2    = float(split_payload[5])
                current2    = float(split_payload[6])
                power2      = float(split_payload[7])
                relay1      = int(split_payload[8])
                relay2      = int(split_payload[9])
                relay3      = int(split_payload[10])
                relay4      = int(split_payload[11])

                # Insert into SQLite
                query = """
                    INSERT INTO four_channel (
                        device_type, device_id, 
                        voltage, current, power, 
                        inserttimestamp, 
                        relay1, relay2, relay3, relay4,
                        voltage2, current2, power2
                    ) 
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """
                val = (
                    "BTB4Channel", device_id,
                    voltage1, current1, power1,
                    current_time,
                    relay1, relay2, relay3, relay4,
                    voltage2, current2, power2
                )
                cursor.execute(query, val)
                connection.commit()

                # Insert into Excel
                excel_row = [
                    "BTB4Channel", device_id,
                    voltage1, current1, power1,
                    voltage2, current2, power2,
                    relay1, relay2, relay3, relay4,
                    current_time
                ]
                insert_excel_row(excel_row)

                print(f"✅ Data inserted for device {device_id}")

        connection.close()

    except Exception as e:
        print(f"❌ Error processing message: {e}")

# -------------------
# MQTT Start
# -------------------
mqttc.on_connect = on_connect
mqttc.on_message = on_message
mqttc.on_disconnect = on_disconnect

def mqtt_connect():
    try:
        print("🔌 Connecting to AWS IoT Core...")
        mqttc.connect(broker, port, keepalive=60)
        mqttc.loop_start()
    except Exception as e:
        print(f"❌ MQTT connection error: {e}")

# -------------------
# Main
# -------------------
if __name__ == "__main__":
    init_db()
    init_excel()
    mqtt_connect()
    print("🚀 System running...")

    # Keep alive
    while True:
        time.sleep(1)
